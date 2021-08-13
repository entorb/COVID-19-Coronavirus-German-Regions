#!/usr/bin/python3
# -*- coding: utf-8 -*-

"""
fetches mortality data from Destatis
see https://www.destatis.de/DE/Themen/Querschnitt/Corona/Gesellschaft/bevoelkerung-sterbefaelle.html
data: https://www.destatis.de/DE/Themen/Gesellschaft-Umwelt/Bevoelkerung/Sterbefaelle-Lebenserwartung/Tabellen/sonderauswertung-sterbefaelle.html;jsessionid=3B59CB1FA0C08C059243535606A41FBF.internet8721
"""

__author__ = "Dr. Torben Menke"
__email__ = "https://entorb.net"
__license__ = "GPL"

# Built-in/Generic Imports
# import codecs
# import urllib
# import requests
# import csv
# import json
# import requests
import os

import pandas as pd
import openpyxl
import csv
import urllib.request

# my helper modules
import helper

# 1. read my covid data, drop starting empty 2 days from feb, prepend 31+28=59 empty rows for matching

# Jan und Feb values are missing for Covid Deaths series, so I need a couple of empty rows
l = [None] * 59
df1 = pd.DataFrame(data={'Deaths_Covid_2020': l})

df0 = pd.read_csv('data/de-states/de-state-DE-total.tsv', sep="\t")
df2 = pd.DataFrame()
df2['Date'] = df0['Date']
df2['Deaths_Covid_2020'] = df0['Deaths_New']
del df0

# ensure first row is from 28.2
assert (df2.iloc[0]['Date'] ==
        '2020-02-28'), "Error of start date, expecting 2020-02-28"
# drop the 2 Feb data rows (of 0 deaths)
df2.drop([0, 1], inplace=True)

df_covid_2020 = pd.DataFrame()
df_covid_2020['Deaths_Covid_2020'] = df1['Deaths_Covid_2020'].append(
    df2['Deaths_Covid_2020'], ignore_index=True)
df_covid_2020['Deaths_Covid_2020_roll'] = df_covid_2020['Deaths_Covid_2020'].rolling(
    window=7, min_periods=1).mean().round(1)
# print(df_covid_2020.tail())
del df1, df2


# 2. fetch and parse Excel of mortality data from Destatis

excelFile = 'cache\de-mortality.xlsx'


if not helper.check_cache_file_available_and_recent(fname=excelFile, max_age=1800, verbose=False):
    url = "https://www.destatis.de/DE/Themen/Gesellschaft-Umwelt/Bevoelkerung/Sterbefaelle-Lebenserwartung/Tabellen/sonderauswertung-sterbefaelle.xlsx?__blob=publicationFile"
    filedata = urllib.request.urlopen(url)
    datatowrite = filedata.read()
    with open(excelFile, mode='wb') as f:
        f.write(datatowrite)


# data_only : read values instead of formulas
workbookIn = openpyxl.load_workbook(excelFile, data_only=True)
sheetIn = workbookIn['D_2016_2020_Tage']

l_dates = []
l_deaths2020 = []
l_deaths2019 = []
l_deaths2018 = []
l_deaths2017 = []
l_deaths2016 = []
for col in range(2, 368):
    day = sheetIn.cell(column=col, row=9).value
    # we skip the 29.02. for each year
    if day == '29.02.':
        continue
    l_dates.append(day)
    l_deaths2020.append(sheetIn.cell(column=col, row=10).value)
    l_deaths2019.append(sheetIn.cell(column=col, row=11).value)
    l_deaths2018.append(sheetIn.cell(column=col, row=12).value)
    l_deaths2017.append(sheetIn.cell(column=col, row=13).value)
    l_deaths2016.append(sheetIn.cell(column=col, row=14).value)


data = zip(l_dates, l_deaths2016, l_deaths2017,
           l_deaths2018, l_deaths2019, l_deaths2020)

df = pd.DataFrame(data, columns=['Day', '2016', '2017',
                                 '2018', '2019', '2020'])

df['2016_roll'] = df['2016'].rolling(window=7, min_periods=1).mean().round(1)
df['2017_roll'] = df['2017'].rolling(window=7, min_periods=1).mean().round(1)
df['2018_roll'] = df['2018'].rolling(window=7, min_periods=1).mean().round(1)
df['2019_roll'] = df['2019'].rolling(window=7, min_periods=1).mean().round(1)
df['2020_roll'] = df['2020'].rolling(window=7, min_periods=1).mean().round(1)
df['2016_2019_mean'] = df.iloc[:, [1, 2, 3, 4]
                               ].mean(axis=1)  # not column 0 = day
df['2016_2019_mean_roll'] = df['2016_2019_mean'].rolling(
    window=7, min_periods=1).mean().round(1)

df['2016_2019_roll_max'] = df.iloc[:, [6, 7, 8, 9]].max(axis=1)
df['2016_2019_roll_min'] = df.iloc[:, [6, 7, 8, 9]].min(axis=1)

df = df.join(df_covid_2020)

df.to_csv('data/de-mortality.tsv', sep="\t", index=False)
