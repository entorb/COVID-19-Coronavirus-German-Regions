#!/usr/bin/env python3.10
# by Dr. Torben Menke https://entorb.net
# https://github.com/entorb/COVID-19-Coronavirus-German-Regions
"""
fetches mortality data from Destatis
see https://www.destatis.de/DE/Themen/Querschnitt/Corona/Gesellschaft/bevoelkerung-sterbefaelle.html
data: https://www.destatis.de/DE/Themen/Gesellschaft-Umwelt/Bevoelkerung/Sterbefaelle-Lebenserwartung/Tabellen/sonderauswertung-sterbefaelle.html;jsessionid=3B59CB1FA0C08C059243535606A41FBF.internet8721
"""
import datetime as dt

import openpyxl
import pandas as pd

import helper


# 1. read and prepare my covid data


def prepare_covid_data() -> pd.DataFrame:
    """
    read my covid deaths data
    remove 29.2. (Schaltjahre)
    returns a df having the day as index in format "dd.mm."
    """
    df = pd.read_csv(
        "data/de-states/de-state-DE-total.tsv",
        sep="\t",
        usecols=[
            "Date",
            "Deaths_New",
        ],
        parse_dates=[
            "Date",
        ],
        index_col="Date",
    )

    df = df.rename(columns={"Deaths_New": "Deaths_Covid"}, errors="raise")

    assert df.index[0] == pd.to_datetime(
        "2020-01-02",
    ), f"Error of start date, expecting 2020-01-02, got : {df.index[0]}"

    # add dummy row for missing 1.1.2020
    df.loc[pd.to_datetime("2020-01-01")] = 0
    df = df.sort_index()  # sorting by index

    # drop deaths of last 4 weeks, as they are not final yet
    date_4w = dt.date.today() - dt.timedelta(weeks=4)
    df = df[df.index < pd.to_datetime(date_4w)]
    del date_4w

    # add day in format 01.03.
    df["Day"] = df.index.strftime("%d.%m.")

    # rolling average
    df = helper.pandas_calc_roll_av(df=df, column="Deaths_Covid", days=7)

    # remove 29.2. (Schaltjahre) (after calc of rolling av)
    df = df[df["Day"] != "29.02."]

    # one df per year
    df_covid_2020 = (
        df[df.index.year == 2020][["Day", "Deaths_Covid", "Deaths_Covid_roll_av"]]
        .reset_index(drop=True)
        .rename(
            columns={
                "Deaths_Covid": "Deaths_Covid_2020",
                "Deaths_Covid_roll_av": "Deaths_Covid_2020_roll_av",
            },
            errors="raise",
        )
    )
    df_covid_2020.set_index("Day", inplace=True)

    df_covid_2021 = (
        df[df.index.year == 2021][["Day", "Deaths_Covid", "Deaths_Covid_roll_av"]]
        .reset_index(drop=True)
        .rename(
            columns={
                "Deaths_Covid": "Deaths_Covid_2021",
                "Deaths_Covid_roll_av": "Deaths_Covid_2021_roll_av",
            },
            errors="raise",
        )
    )
    df_covid_2021.set_index("Day", inplace=True)

    df_covid_2022 = (
        df[df.index.year == 2022][["Day", "Deaths_Covid", "Deaths_Covid_roll_av"]]
        .reset_index(drop=True)
        .rename(
            columns={
                "Deaths_Covid": "Deaths_Covid_2022",
                "Deaths_Covid_roll_av": "Deaths_Covid_2022_roll_av",
            },
            errors="raise",
        )
    )
    df_covid_2022.set_index("Day", inplace=True)

    # join in index = Day
    df_covid = df_covid_2020.join(df_covid_2021).join(df_covid_2022)

    return df_covid


def convert2date(year: int, ddmm: str) -> dt.date:
    # test: print(convert2date(year=2016, ddmm="01.01"))
    d, m, empty = ddmm.split(".")
    date = dt.date(int(year), int(m), int(d))
    return date


def fetch_and_prepare_mortality_data_timeseries() -> pd.DataFrame:
    """
    fetch and parse Excel of mortality data from Destatis
    """

    excelFile = "cache/de-mortality.xlsx"

    # as file is stored in cache folder which is not part of the commit, we can use the caching here
    helper.download_from_url_if_old(
        url="https://www.destatis.de/DE/Themen/Gesellschaft-Umwelt/Bevoelkerung/Sterbefaelle-Lebenserwartung/Tabellen/sonderauswertung-sterbefaelle.xlsx?__blob=publicationFile",
        # url="https://www.destatis.de/DE/Themen/Gesellschaft-Umwelt/Bevoelkerung/Sterbefaelle-Lebenserwartung/Tabellen/sonderauswertung-sterbefaelle.xlsx;jsessionid=FB723BC229CAC6B6302FF752CC66DE7C.live742?__blob=publicationFile",
        file_local=excelFile,
        max_age=3600,
        verbose=True,
    )

    # if not helper.check_cache_file_available_and_recent(
    #     fname=excelFile,
    #     max_age=1800,
    #     verbose=False,  # as file is stored in cache folder which is not part of the commit, we can use the caching here
    # ):
    #     url = "https://www.destatis.de/DE/Themen/Gesellschaft-Umwelt/Bevoelkerung/Sterbefaelle-Lebenserwartung/Tabellen/sonderauswertung-sterbefaelle.xlsx?__blob=publicationFile"
    #     filedata = urllib.request.urlopen(url)
    #     datatowrite = filedata.read()
    #     with open(excelFile, mode="wb") as f:
    #         f.write(datatowrite)

    # data_only : read values instead of formulas
    workbookIn = openpyxl.load_workbook(excelFile, data_only=True)
    sheetIn = workbookIn["D_2016_2022_Tage"]

    # 1. time series for correct rolling av calc

    # 1.1 read from Excel
    l_timeseries = []
    col = 1
    for row in range(16, 10 - 1, -1):
        year = int(sheetIn.cell(column=col, row=row).value)
        assert year >= 2016
        assert year <= 2028
        for col2 in range(2, 368):
            day_str = sheetIn.cell(column=col2, row=9).value
            deaths = sheetIn.cell(column=col2, row=row).value
            if deaths in (
                "X",  # 29.2. and not "Schaltjahr"
                None,  # blank
            ):
                continue
            # print(f"'{value}'")
            date = convert2date(year=year, ddmm=day_str)
            l_timeseries.append((date, deaths))  # day_str = dd.mm.

    df = pd.DataFrame(
        data=l_timeseries,
        columns=[
            "Date",
            "Deaths",
        ],
    )
    df = helper.pandas_set_date_index(df=df, date_column="Date")

    # assert that we start with 1.1.2016
    assert df.index[0] == pd.to_datetime("2016-01-01")

    df = helper.pandas_calc_roll_av(df=df, column="Deaths", days=7)

    df[["Deaths", "Deaths_roll_av"]].to_csv(
        "data/ts-de-mortality.tsv",
        sep="\t",
        line_terminator="\n",
    )
    return df


def merge_mortality_data_per_day(df: pd.DataFrame) -> pd.DataFrame:
    """
    convert mortalitiy timeseries to format: day_str , 2016, 2017,...
    """

    # add day column in format 01.03.
    df["Day"] = df.index.strftime("%d.%m.")

    # remove 29.2. (Schaltjahre) (after calc of rolling av in timeseries)
    df = df[~(df["Day"] == "29.02.")]

    l_days = df[df.index.year == 2016]["Day"].tolist()
    assert len(l_days) == 365

    # create new empty df
    df2 = pd.DataFrame()
    # (index=l_days, data={})

    # add full year data of columns Deaths and Deaths_roll_av
    for year in range(2016, 2021 + 1, 1):
        df2[str(year)] = df[df.index.year == year]["Deaths"].tolist()
        df2[str(year) + "_roll_av"] = df[df.index.year == year][
            "Deaths_roll_av"
        ].tolist()
    del year

    # add current year
    df2["2022"] = pd.Series(df[df.index.year == 2022]["Deaths"].tolist())
    df2["2022_roll_av"] = pd.Series(
        df[df.index.year == 2022]["Deaths_roll_av"].tolist(),
    )

    # setting the index to the Day
    df2.index = l_days
    df2.index.name = "Day"

    # calculations
    df2["2016_2019_mean"] = df2[["2016", "2017", "2018", "2019"]].mean(axis=1)

    df2 = helper.pandas_calc_roll_av(df=df2, column="2016_2019_mean", days=7)

    df2["2016_2019_roll_av_min"] = df2[
        ["2016_roll_av", "2017_roll_av", "2018_roll_av", "2019_roll_av"]
    ].min(axis=1)

    df2["2016_2019_roll_av_max"] = df2[
        ["2016_roll_av", "2017_roll_av", "2018_roll_av", "2019_roll_av"]
    ].max(axis=1)

    return df2


if __name__ == "__main__":
    df_covid = prepare_covid_data()

    df_mortality_ts = fetch_and_prepare_mortality_data_timeseries()
    df_mortality = merge_mortality_data_per_day(df_mortality_ts)

    df = df_mortality.join(df_covid)

    df.to_csv("data/de-mortality.tsv", sep="\t", index=True, line_terminator="\n")
